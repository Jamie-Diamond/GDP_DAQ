from Data_import import PP_data_import
from Plotting_ToolBox import linar_var_plot, GPS_plot
import matplotlib.pyplot


def polarFilter(Data, angleRange):
    index = 0
    polarPoints = []
    for i in Data:
        if index > 3 and index < len(Data) - 2:
            if abs(i[1]["HDG"] - Data[index - 2][1]["HDG"]) < angleRange \
            and abs(i[1]["HDG"] - Data[index - 1][1]["HDG"]) < angleRange \
            and abs(i[1]["HDG"] - Data[index + 1][1]["HDG"]) < angleRange \
            and abs(i[1]["HDG"] - Data[index + 2][1]["HDG"]) < angleRange:
                #print(abs(i[1]["HDG"] - Data[index - 2][1]["HDG"]), abs(i[1]["HDG"] - Data[index - 1][1]["HDG"]), abs(i[1]["HDG"] - Data[index + 1][1]["HDG"]), abs(i[1]["HDG"] - Data[index + 2][1]["HDG"]))
                polarPoints.append(i)

        index += 1

    return polarPoints


def plotPolars(Data, windSpeed=15, WindTol=15, anglerange=15, minspeed=5, bodge=0):
    Data = polarFilter(Data, anglerange)
    import math
    from Data_import import wraptopm180
    theta = []
    r = []
    for i in Data:
        if i[1]["TWS"] is not None:
            var = abs(i[1]["TWS"] - windSpeed)
            if var < WindTol:
                if i[1]["BSP"] > minspeed:
                    theta.append(math.radians(wraptopm180(i[1]["TWA"]+bodge)))
                    r.append(i[1]["BSP"])
    matplotlib.pyplot.figure(101)
    axis = matplotlib.pyplot.subplot(111, projection='polar')
    axis.set_theta_zero_location('N')
    axis.set_theta_direction(-1)
    axis.set_rlabel_position(math.pi/2)
    axis.set_rlim(0, 18)
    axis.plot(theta, r, '.b')
    axis.grid(True)
    axis.set_title("BSP vs TWA (TWS: " + str(windSpeed) + " ± " + str(WindTol) + '  HDG Filter: ± '+str(anglerange) + '  Min Speed =  '+str(minspeed) + ' Bodged by:' +str(bodge)+'º' + ')', va='bottom')

def plotAngleAveragedPolar(Data, windSpeed=15, WindTol=15, anglerange=15, minspeed=5, averange=0.5, bodge=0, mirror=False, fig=1):
    Data = polarFilter(Data, anglerange)
    import math
    import heapq
    from numpy import arange, mean
    from Data_import import roundNo, Wrapto0_360, Wrapto180

    r = []
    theta = []
    angles = {}
    for i in arange(0, 360, averange):
        angles[i] = []

    for i in Data:
        if i[1]["TWS"] is not None:
            var = abs(i[1]["TWS"] - windSpeed)
            if var < WindTol:
                if i[1]["BSP"] > minspeed:
                    if not mirror:
                        roundedAngle = roundNo(Wrapto0_360(i[1]["TWA"]+bodge), averange)
                        angles[roundedAngle].append(i[1]["BSP"])
                    else:
                        roundedAngle = roundNo(Wrapto180(i[1]["TWA"] + bodge), averange)
                        angles[roundedAngle].append(i[1]["BSP"])
    for i in angles:
        if len(angles[i]) != 0:
            angles[i] = mean(heapq.nlargest(8, angles[i]))
            theta.append(math.radians(i))
            r.append(angles[i])
        else:
            angles[i] = None
    print(angles)
    matplotlib.pyplot.figure(fig)
    axis = matplotlib.pyplot.subplot(111, projection='polar')
    axis.set_theta_zero_location('N')
    axis.set_theta_direction(-1)
    axis.set_rlabel_position(math.pi / 2)
    axis.set_rlim(0, 18)
    axis.plot(theta, r, '.b')
    axis.grid(True)
    axis.set_title("BSP vs TWA (TWS: " + str(windSpeed) + "±" + str(WindTol) + ', HDG Filter: ±' + str(
        anglerange) + ', Min Speed: ' + str(minspeed) + "" + ', Angle Mapping: ±' + str(averange) + ")", va='bottom')


def VPP_plot_from_excel():
    import openpyxl
    import math
    import matplotlib
    wb = openpyxl.load_workbook('F18 SailData.xlsx')
    sheet = wb.get_sheet_by_name('Polars 06-12 Valid')
    speeds12 = []
    speeds14 = []
    angles = []
    for i in range(5, 21, 1):
        speeds12.append(sheet.cell(row=i, column=9).value)
        speeds14.append(sheet.cell(row=i, column=8).value)
        angles.append(math.radians(sheet.cell(row=i, column=1).value))
    axis = matplotlib.pyplot.subplot(111, projection='polar')
    axis.set_theta_zero_location('N')
    axis.set_theta_direction(-1)
    axis.set_rlabel_position(math.pi / 2)
    axis.set_rlim(0, 18)
    axis.plot(angles, speeds12, '.r-')
    axis.plot(angles, speeds14, '.k-')
    axis.grid(True)



if __name__ == "__main__":
    data = PP_data_import(reprocess=False)
    plotAngleAveragedPolar(data, windSpeed=13, WindTol=3, anglerange=40, minspeed=6, averange=5, bodge=18, fig=1)
    plotAngleAveragedPolar(data, windSpeed=13, WindTol=3, anglerange=40, minspeed=6, averange=5, bodge=18, mirror=True, fig=2)
    VPP_plot_from_excel()
    matplotlib.pyplot.show()
    #linar_var_plot(data, ['BSP', 'GWS', 'TWS', 'AWS'])
    linar_var_plot(data, ['COG', 'COW', 'HDG', 'LWY'])
    print('Finito')
